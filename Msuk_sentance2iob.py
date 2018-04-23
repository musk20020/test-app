#encoding=utf-8

import pandas as pd
import jieba
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re
import numpy as np

#data = pd.read_excel('Musk_iob_input.xlsx',sheet_name=0)
data = pd.read_excel('primiry130.xls',sheet_name=0)

#library = pd.read_excel('entity_type.xlsx',sheet_name=1)
library = pd.read_csv('entity_type.csv')
wb = Workbook()
ws = wb.active
file = open("word2index_calibrated.txt","r").readlines()
#file2 = open("word2index_musk.txt","w")

file_iob = open("ijia_primiry130_musk.iob","w")
#file_iob = open("ijia_dataset_musk.iob","w")

#sheet_name --> 讀取excel的第幾個分頁
traing_iob = "musk_training_iob.txt"

library_condition = library.iloc[:,[1,2,3,7,8,9,10,11,12,13,14,15,16,17]]
#print(library_condition.values[2][2] == u"咖啡")

'''
print(library.shape[0])
print(library.shape[1])
print(library.values[1606][1])
for i in range(1,library.shape[1]):
    print(i)
    for j in range( 1, library.shape[0] ):
        #print(library.values[j][i])
        keys = library.values[j][i]
        if(keys != u"non" and type(keys) == type(u"unicode")):
            #keys = re.sub(u"\n",u"",keys)
            keys = keys.encode("utf-8") + "\n"
            if(keys not in file):
                file2.write(keys)
        else:
            break
'''

def add_road_dic():
    file = "./road_dic/road"
    rows = 1
    keys = []
    for i in range(1,42):
        road = open(file+str(i)+".txt","r").readlines()
        road = re.sub("１","一",road[0])
        road = re.sub("２","二",road)
        road = re.sub("３","三",road)
        road = re.sub("４","四",road)
        road = re.sub("５","五",road)
        road = re.sub("６","六",road)
        road = re.sub("７","七",road)
        road = re.sub("８","八",road)
        road = re.sub("９","九",road).split(",")
        for i in range(0,len(road)):
            keys.append(road[i])

    for j in range( 1, library.shape[0] ):
        # print(library.values[j][i])
        keys.append(library.values[j][7])
        if j%1000 == 0:
            print("已處理"+str(j)+"筆")
    road = set(keys)
    for r in road:
        ws.cell( row=rows, column=1 ).value = r
        rows += 1
        if rows % 1000 == 0:
            print("已處理" + str(j) + "筆")
    wb.save( 'new_road.xlsx' )

#add_road_dic()


jieba.set_dictionary
jieba.load_userdict(file)

for i in range(0,data.shape[0]):
    utterance = data.values[i][0].encode("utf-8")
    word = jieba.cut( str( utterance ) )
    text = " ".join( word )
    words = text.split(" ")
    text_out = ""
    for j in words:
        index = np.argwhere(library_condition.values[:][:] == j.encode("utf-8"))
        if(index.any()):
            text_out = text_out + " " + library_condition.values[0][index[0][1]]
        else:
            text_out = text_out + " O"
    input_text = "BOS " + text.encode("utf-8") + " EOS\t"
    output_text = "O" + text_out
    intent = utterance = " " + data.values[i][1].encode("utf-8")
    file_iob.write(input_text+output_text+intent+"\n")
    print(input_text+output_text+intent)
